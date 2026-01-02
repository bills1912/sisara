import io
from typing import List
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
# Pastikan import model sesuai dengan struktur project Anda
from app.models.schemas import BudgetRowResponse, RowType 

class ExportService:
    def __init__(self):
        # Mapping indentasi visual berdasarkan tipe
        self.indent_map = {
            RowType.SATKER: 0,
            RowType.PROGRAM: 1,
            RowType.ACTIVITY: 2,
            RowType.KRO: 3,
            RowType.RO: 4,
            RowType.COMPONENT: 5,
            RowType.SUBCOMPONENT: 6,
            RowType.ACCOUNT: 7,
            RowType.DETAIL: 8
        }
        
        # Mapping warna background Excel (Hex ARGB) agar sesuai UI
        self.color_map = {
            RowType.SATKER: "FFF3F4F6",    # gray-100
            RowType.PROGRAM: "FFE0F2FE",   # sky-100
            RowType.ACTIVITY: "FFEEF2FF",  # indigo-50
            RowType.KRO: "FFECFDF5",       # emerald-50
            RowType.RO: "FFFFFBEB",        # amber-50
            RowType.ACCOUNT: "FFF9FAFB",   # gray-50
        }

    def _flatten_tree(self, nodes: List[BudgetRowResponse], result: List[BudgetRowResponse]):
        """Helper rekursif untuk meratakan tree menjadi list"""
        for node in nodes:
            result.append(node)
            if node.children:
                self._flatten_tree(node.children, result)

    def generate_excel(self, data: List[BudgetRowResponse]) -> io.BytesIO:
        wb = Workbook()
        ws = wb.active
        ws.title = "Rincian Kertas Kerja"

        # --- 1. SETUP HEADERS ---
        # Definisi Header Baris 1
        headers_1 = [
            ("Kode", 1, 3), ("Uraian", 1, 3), 
            ("Semula", 4, 1), ("Menjadi", 4, 1), 
            ("Rencana Penarikan Dana (RPD) & Realisasi", 13, 1)
        ]
        
        col_idx = 1
        for title, width, height in headers_1:
            cell = ws.cell(row=1, column=col_idx, value=title)
            if width > 1 or height > 1:
                ws.merge_cells(start_row=1, start_column=col_idx, 
                               end_row=1+(height-1), end_column=col_idx+(width-1))
            col_idx += width

        # Definisi Header Baris 2 (Sub-headers)
        sub_headers = ["Vol", "Sat", "Harga", "Jumlah"]
        
        # Sub-header Semula (Kolom 3-6)
        for i, h in enumerate(sub_headers):
            cell = ws.cell(row=2, column=3+i, value=h)
            ws.merge_cells(start_row=2, start_column=3+i, end_row=3, end_column=3+i)

        # Sub-header Menjadi (Kolom 7-10)
        for i, h in enumerate(sub_headers):
            cell = ws.cell(row=2, column=7+i, value=h)
            ws.merge_cells(start_row=2, start_column=7+i, end_row=3, end_column=7+i)

        # Sub-header Bulan (Kolom 11-23)
        months = ['Jan', 'Feb', 'Mar', 'Apr', 'Mei', 'Jun', 'Jul', 'Agu', 'Sep', 'Okt', 'Nov', 'Des', 'Total']
        for i, m in enumerate(months):
            cell = ws.cell(row=2, column=11+i, value=m)
            ws.merge_cells(start_row=2, start_column=11+i, end_row=3, end_column=11+i)

        # Styling Headers
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="FF1E3A8A", end_color="FF1E3A8A", fill_type="solid") # Blue 800
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        for row in ws.iter_rows(min_row=1, max_row=3, max_col=23):
            for cell in row:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                cell.border = thin_border

        # Adjust Column Widths
        ws.column_dimensions['A'].width = 18  # Kode
        ws.column_dimensions['B'].width = 65  # Uraian (Lebih lebar)
        for i in range(3, 24):
            ws.column_dimensions[get_column_letter(i)].width = 15

        # *** FITUR FREEZE PANES ***
        # Membekukan Baris 1-3 (Header) dan Kolom A-B (Kode & Uraian)
        # 'C4' berarti pane akan beku di sebelah kiri C (A, B) dan di atas 4 (1, 2, 3)
        ws.freeze_panes = "C4"

        # --- 2. POPULATE DATA ---
        flat_data = []
        self._flatten_tree(data, flat_data)

        current_row = 4
        for item in flat_data:
            # Kolom Kode
            ws.cell(row=current_row, column=1, value=item.code).alignment = Alignment(vertical='top')
            
            # Kolom Uraian dengan Indentasi
            desc_cell = ws.cell(row=current_row, column=2, value=item.description)
            indent_level = self.indent_map.get(item.type, 0)
            desc_cell.alignment = Alignment(indent=indent_level, vertical='top', wrap_text=True)

            # Styling Baris (Warna & Bold)
            is_bold = item.type not in [RowType.DETAIL, RowType.SUBCOMPONENT]
            row_fill = None
            
            if item.type in self.color_map:
                c_hex = self.color_map[item.type]
                row_fill = PatternFill(start_color=c_hex, end_color=c_hex, fill_type="solid")

            for col in range(1, 24):
                cell = ws.cell(row=current_row, column=col)
                if row_fill:
                    cell.fill = row_fill
                if is_bold:
                    cell.font = Font(bold=True)
                cell.border = thin_border

            # Data Semula
            if item.semula:
                ws.cell(row=current_row, column=3, value=item.semula.volume)
                ws.cell(row=current_row, column=4, value=item.semula.unit)
                ws.cell(row=current_row, column=5, value=item.semula.price).number_format = '#,##0'
                ws.cell(row=current_row, column=6, value=item.semula.total).number_format = '#,##0'

            # Data Menjadi
            if item.menjadi:
                ws.cell(row=current_row, column=7, value=item.menjadi.volume)
                ws.cell(row=current_row, column=8, value=item.menjadi.unit)
                ws.cell(row=current_row, column=9, value=item.menjadi.price).number_format = '#,##0'
                ws.cell(row=current_row, column=10, value=item.menjadi.total).number_format = '#,##0'

            # Data Bulanan (RPD + Realisasi)
            total_rpd = 0
            if hasattr(item, 'monthlyAllocation') and item.monthlyAllocation:
                for i in range(12):
                    m_key = str(i) # Kunci bulan biasanya string "0", "1", dst.
                    val = 0
                    
                    # Akses data bulanan secara aman (baik dict maupun object)
                    # Sesuaikan dengan cara data Anda di-load di backend
                    allocs = item.monthlyAllocation
                    if isinstance(allocs, dict) and m_key in allocs:
                        detail = allocs[m_key]
                        # Ambil RPD dan Realisasi
                        rpd = detail.rpd if hasattr(detail, 'rpd') else detail.get('rpd', 0)
                        real = detail.realization if hasattr(detail, 'realization') else detail.get('realization', 0)
                        val = rpd + real
                    elif hasattr(allocs, m_key):
                        # Jika Pydantic model akses langsung
                        detail = getattr(allocs, m_key)
                        if detail:
                             val = detail.rpd + detail.realization

                    ws.cell(row=current_row, column=11+i, value=val).number_format = '#,##0'
                    total_rpd += val

            # Total Tahunan
            ws.cell(row=current_row, column=23, value=total_rpd).number_format = '#,##0'

            current_row += 1

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output

export_service = ExportService()